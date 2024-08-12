import os
import time
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

import torch
import torch.nn as nn
import torch.optim as optim
from model.Ynet import losses
from model.Ynet.losses import CharbonnierLoss
# from model.YNet.dataset_utils import MixUp_AUG
from prep import printProgressBar
import openpyxl


# from thop import profile
from measure import compute_measure

def _tensor_size(t):
    return t.size()[1] * t.size()[2] * t.size()[3]

def count_parameters(model):
    return sum(p.numel() for p in model.parameters() if p.requires_grad)

def split_arr(arr,patch_size,stride=32):    ## 512*512 to 32*32
    pad = (16, 16, 16, 16) # pad by (0, 1), (2, 1), and (3, 3)
    arr = nn.functional.pad(arr, pad, "constant", 0)
    _,_,h,w = arr.shape
    num = h//stride - 1
    arrs = torch.zeros(num*num,1,patch_size,patch_size)

    for i in range(num):
        for j in range(num):
            arrs[i*num+j,0] = arr[0,0,i*stride:i*stride+patch_size,j*stride:j*stride+patch_size]
    return arrs

def agg_arr(arrs, size, stride=32):  ## from 32*32 to size 512*512
    arr = torch.zeros(size, size)
    n,_,h,w = arrs.shape
    num = size//stride
    for i in range(num):
        for j in range(num):
            arr[i*stride:(i+1)*stride,j*stride:(j+1)*stride] = arrs[i*num+j,:,16:48,16:48]
  #return arr
    return arr.unsqueeze(0).unsqueeze(1)

######### Loss ###########
criterion1 = losses.CharbonnierLoss()
criterion2 = losses.EdgeLoss()
# mixup = MixUp_AUG()
class Solver(object):
    def __init__(self, args, data_loader, val_data_loader, model):

        #新增成员变量
        # 设定模型名称
        self.model_name = 'YNet'
        # 设定开始验证和保存模型的epoch
        self.val_epoch = args.val_epoch
        self.val_epoch = args.val_epoch
        # 定义用于验证的dataloader
        self.val_data_loader = val_data_loader
        # 用于测试的epoch
        self.test_epoch = args.test_epoch

        # --------------------------

        self.mode = args.mode
        self.load_mode = args.load_mode
        self.data_loader = data_loader

        if args.device:
            self.device = torch.device(args.device)
        else:
            print('torch.cuda.is_available()', torch.cuda.is_available())
            self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

        self.norm_range_min = args.norm_range_min
        self.norm_range_max = args.norm_range_max
        self.trunc_min = args.trunc_min
        self.trunc_max = args.trunc_max

        self.save_path = args.save_path
        self.multi_gpu = args.multi_gpu

        self.num_epochs = args.num_epochs
        self.print_iters = args.print_iters
        self.decay_iters = args.decay_iters
        self.save_iters = args.save_iters
        self.test_iters = args.test_iters
        self.result_fig = args.result_fig

        self.patch_size = args.patch_size

        self.MODEL = model
        if (self.multi_gpu) and (torch.cuda.device_count() > 1):
            print('Use {} GPUs'.format(torch.cuda.device_count()))
            self.MODEL = nn.DataParallel(self.MODEL, device_ids=[0, 1])
            self.device = torch.device("cuda:0")## data parallel  ,device_ids=[0,1]
        self.MODEL.to(self.device)

        self.lr = args.lr
        # self.criterion = nn.MSELoss()  #nn.L1Loss
        # self.criterion1 = nn.L1Loss()
        # self.criterion = CharbonnierLoss()
        # YNet
        self.criterion1 = CharbonnierLoss()

        self.optimizer = optim.Adam(self.MODEL.parameters(), self.lr)


    def save_model(self, iter_):
        f = os.path.join(self.save_path, 'T2T_vit_{}iter.ckpt'.format(iter_))
        torch.save(self.MODEL.state_dict(), f)

    def save_model_by_epoch(self, epoch_):
        f = os.path.join(self.save_path, self.model_name + '_{}epoch.ckpt'.format(epoch_))
        torch.save(self.MODEL.state_dict(), f)


    def load_model(self, iter_):
        device = torch.device('cpu')
        f = os.path.join(self.save_path, 'T2T_vit_{}iter.ckpt'.format(iter_))
        #f = os.path.join('model_pretrained', 'T2T_vit_{}iter.ckpt'.format(iter_))
        self.MODEL.load_state_dict(torch.load(f, map_location=device), False)

    def load_model_by_epoch(self, epoch_):
        device = torch.device('cpu')
        f = os.path.join(self.save_path, self.model_name + '_{}epoch.ckpt'.format(epoch_))
        #f = os.path.join('model_pretrained', 'T2T_vit_{}iter.ckpt'.format(iter_))
        self.MODEL.load_state_dict(torch.load(f, map_location=device), False)


    def lr_decay(self):
        lr = self.lr * 0.5
        for param_group in self.optimizer.param_groups:
            param_group['lr'] = lr


    def denormalize_(self, image):
        image = image * (self.norm_range_max - self.norm_range_min) + self.norm_range_min
        return image


    def trunc(self, mat):
        mat[mat <= self.trunc_min] = self.trunc_min
        mat[mat >= self.trunc_max] = self.trunc_max
        return mat


    def save_fig(self, x, y, pred, fig_name, original_result, pred_result):
        x, y, pred = x.numpy(), y.numpy(), pred.numpy()
        f, ax = plt.subplots(1, 3, figsize=(30, 10))
        ax[0].imshow(x, cmap=plt.cm.gray, vmin=self.trunc_min, vmax=self.trunc_max)
        ax[0].set_title('Quarter-dose', fontsize=30)
        ax[0].set_xlabel("PSNR: {:.4f}\nSSIM: {:.4f}\nRMSE: {:.4f}".format(original_result[0],
                                                                           original_result[1],
                                                                           original_result[2]), fontsize=20)
        ax[1].imshow(pred, cmap=plt.cm.gray, vmin=self.trunc_min, vmax=self.trunc_max)
        ax[1].set_title('Result', fontsize=30)
        ax[1].set_xlabel("PSNR: {:.4f}\nSSIM: {:.4f}\nRMSE: {:.4f}".format(pred_result[0],
                                                                           pred_result[1],
                                                                           pred_result[2]), fontsize=20)
        ax[2].imshow(y, cmap=plt.cm.gray, vmin=self.trunc_min, vmax=self.trunc_max)
        ax[2].set_title('Full-dose', fontsize=30)

        f.savefig(os.path.join(self.save_path, 'fig', 'result_{}.png'.format(fig_name)))
        plt.close()


    def train(self):
        NumOfParam = count_parameters(self.MODEL)
        print('trainable parameter:', NumOfParam)
        #
        # dummy_input = torch.randn(16, 1, 64, 64).to(self.device)
        # flops, params = profile(self.MODEL.module, inputs=(dummy_input,))
        # print('trainable flops:', flops)
        # print('trainable parameter:', params)


        # 加载xlsx文件
        wb = openpyxl.load_workbook(self.model_name + '.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')
        sheet.append(['EPOCH', 'PSNR', 'SSIM', 'RMSE'])
        wb.save(self.model_name + '.xlsx')


        train_losses = []
        total_iters = 0

        start_time = time.time()
        loss_all = []
        loss_part = []
        for epoch in range(1, self.num_epochs+1):
            self.MODEL.train(True)
            for iter_, (x, y) in enumerate(self.data_loader):

                total_iters += 1

                # add 1 channel
                x = x.unsqueeze(0).float().to(self.device)   ## expand one dimension given the dimension 0  4->[1,4]
                y = y.unsqueeze(0).float().to(self.device)   ## copy data to device


                if self.patch_size: # patch training
                    x = x.view(-1, 1, self.patch_size, self.patch_size)  ## similar to reshape
                    y = y.view(-1, 1, self.patch_size, self.patch_size)

                # print(x.shape)
                #!!! pred = self.MODEL(x)
                # pred, noise_level = self.MODEL(x)
                pred = self.MODEL(x)
                # print("pred:", pred.shape)
                # print("noise_level:", noise_level.shape)

                # h_x = noise_level.size()[2]
                # w_x = noise_level.size()[3]
                # count_h = _tensor_size(noise_level[:, :, 1:, :])
                # count_w = _tensor_size(noise_level[:, :, :, 1:])
                #
                # h_tv = torch.pow((noise_level[:, :, 1:, :] - noise_level[:, :, :h_x - 1, :]), 2).sum()
                # w_tv = torch.pow((noise_level[:, :, :, 1:] - noise_level[:, :, :, :w_x - 1]), 2).sum()
                # tvloss = h_tv / count_h + w_tv / count_w

                # loss = self.criterion(pred, y)*100 + 1e-4  ## to prevent 0
                # criterion1 = self.criterion1
                # criterion2 = self.criterion2

                # target = data[0].cuda()
                # input_ = data[1].cuda()
                loss = self.criterion1(pred, y) * 100 + 1e-4  ## to prevent 0
                # loss = criterion1(torch.clamp(pred, 0, 1), y) + 0.1 * criterion2(torch.clamp(pred, 0, 1), y) + 0.05 * tvloss
                #
                # if epoch > 5:
                #     y, x = mixup.aug(y, x)
                # restored = model_restoration(x)
                # loss = criterion1(torch.clamp(pred, 0, 1), y) + 0.1 * criterion2(torch.clamp(pred, 0, 1),
                #                                                                           y)

                # loss = criterion1(torch.clamp(restored, 0, 1), target) + 0.1 * criterion2(torch.clamp(restored, 0, 1),
                #                                                                           target) + 0.05 * tvloss

                # print(loss)
                self.MODEL.zero_grad()
                self.optimizer.zero_grad()

                loss.backward()
                self.optimizer.step()
                train_losses.append(loss.item())

                loss_all.append(loss.item())
                loss_part = loss_all[1:1000]


                # print
                if total_iters % self.print_iters == 0:
                    print("STEP [{}], EPOCH [{}/{}], ITER [{}/{}] \nLOSS: {:.8f}, TIME: {:.1f}s".format(total_iters, epoch, 
                                                                                                        self.num_epochs, iter_+1, 
                                                                                                        len(self.data_loader), loss.item(),
                                                                                                    time.time() - start_time))
                # learning rate decay
                #print(total_iters)  
                if total_iters % self.decay_iters == 0:
                    self.lr_decay()
                # save model
                # if total_iters % self.save_iters == 0:
                #     print("save model: ",total_iters)
                #     self.save_model(total_iters)
                #     np.save(os.path.join(self.save_path, 'loss_{}_iter.npy'.format(total_iters)), np.array(train_losses))

            # 验证
            # 设置开始验证并保存模型的epoch
            if epoch >= self.val_epoch:
            # if epoch % 5 = 0:
                self.val(epoch)
                self.save_model_by_epoch(epoch)

        # self.save_model(total_iters)
        np.save(os.path.join(self.save_path, 'loss_{}_iter.npy'.format(total_iters)), np.array(train_losses))
        print("total_iters:",total_iters)
        ## save loss figure
        plt.plot(np.array(loss_all), 'r')  ## print out the loss curve
        plt.ylim(0, 6)  # 设置纵坐标轴的范围为 0 到 6
        plt.xlabel('Iteration')
        plt.ylabel('Loss')
        plt.title('Loss Curve')
        #plt.show()
        plt.savefig('save/loss.png')

        # 记录训练总时长
        finish_time = time.time() - start_time
        wb = openpyxl.load_workbook(self.model_name + '.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')
        sheet.append(['训练总时长为：', finish_time])
        wb.save(self.model_name + '.xlsx')

    def val(self, val_epoch):
        # del self.MODEL
        # self.MODEL = CTformer(img_size=64,tokens_type='performer', embed_dim=64, depth=1, num_heads=8, kernel=8, stride=4, mlp_ratio=2., token_dim=64)
        # self.CTFormer = T2T_ViT(img_size=128,tokens_type='convolution', in_chans=8,embed_dim=768, depth=6, num_heads=12, kernel=16, stride=8, mlp_ratio=2.)
        # if (self.multi_gpu) and (torch.cuda.device_count() > 1):
        # print('Use {} GPUs'.format(torch.cuda.device_count()))
        # self.MODEL = nn.DataParallel(self.MODEL)   ## data parallel
        self.MODEL.to(self.device)

        # compute PSNR, SSIM, RMSE
        ori_psnr_avg, ori_ssim_avg, ori_rmse_avg = 0, 0, 0
        pred_psnr_avg, pred_ssim_avg, pred_rmse_avg = 0, 0, 0

        ori_psnr_max, ori_ssim_max, ori_rmse_max = 0, 0, 0
        ori_psnr_min, ori_ssim_min, ori_rmse_min = 100, 100, 100

        pred_psnr_max, pred_ssim_max, pred_rmse_max = 0, 0, 0
        pred_psnr_min, pred_ssim_min, pred_rmse_min = 100, 100, 100

        with torch.no_grad():
            for i, (x, y) in enumerate(self.val_data_loader):
                # print(x.shape)
                # print(y.shape)
                shape_ = x.shape[-1]
                x = x.unsqueeze(0).float().to(self.device)
                y = y.unsqueeze(0).float().to(self.device)

                arrs = split_arr(x, 64).to(self.device)  ## split to image patches for test into 4 patches

                # arrs = torch.as_tensor(arrs, dtype=torch.float32)
                # print(arrs)

                arrs[0:64] = self.MODEL(arrs[0:64].cuda())
                arrs[64:2 * 64] = self.MODEL(arrs[64:2 * 64].cuda())
                arrs[2 * 64:3 * 64] = self.MODEL(arrs[2 * 64:3 * 64].cuda())
                arrs[3 * 64:4 * 64] = self.MODEL(arrs[3 * 64:4 * 64].cuda())
                pred = agg_arr(arrs, 512).to(self.device)

                # result = self.MODEL(arrs[0:64])
                # arrs[0:64] = result[0]ITjiangta
                # result1 = self.MODEL(arrs[64:2 * 64])
                # arrs[64:2 * 64] = result1[0]
                # result2 = self.MODEL(arrs[2 * 64:3 * 64])
                # arrs[2 * 64:3 * 64] = result2[0]
                # result3 = self.MODEL(arrs[3 * 64:4 * 64])
                # arrs[3 * 64:4 * 64] = result3[0]
                # pred = agg_arr(arrs, 512).to(self.device)

                # pred = x - pred# denormalize, truncate
                x = self.trunc(self.denormalize_(x.view(shape_, shape_).cpu().detach()))
                y = self.trunc(self.denormalize_(y.view(shape_, shape_).cpu().detach()))
                pred = self.trunc(self.denormalize_(pred.view(shape_, shape_).cpu().detach()))

                data_range = self.trunc_max - self.trunc_min

                original_result, pred_result = compute_measure(x, y, pred, data_range)
                ori_psnr_avg += original_result[0]
                if original_result[0] > ori_psnr_max:
                    ori_psnr_max = original_result[0]
                if original_result[0] < ori_psnr_min:
                    ori_psnr_min = original_result[0]

                ori_ssim_avg += original_result[1]
                if original_result[1] > ori_ssim_max:
                    ori_ssim_max = original_result[1]
                if original_result[1] < ori_ssim_min:
                    ori_ssim_min = original_result[1]

                ori_rmse_avg += original_result[2]
                if original_result[2] > ori_rmse_max:
                    ori_rmse_max = original_result[2]
                if original_result[2] < ori_rmse_min:
                    ori_rmse_min = original_result[2]

                pred_psnr_avg += pred_result[0]
                if pred_result[0] > pred_psnr_max:
                    pred_psnr_max = pred_result[0]
                if pred_result[0] < pred_psnr_min:
                    pred_psnr_min = pred_result[0]

                pred_ssim_avg += pred_result[1]
                if pred_result[1] > pred_ssim_max:
                    pred_ssim_max = pred_result[1]
                if pred_result[1] < pred_ssim_min:
                    pred_ssim_min = pred_result[1]

                pred_rmse_avg += pred_result[2]
                if pred_result[2] > pred_rmse_max:
                    pred_rmse_max = pred_result[2]
                if pred_result[2] < pred_rmse_min:
                    pred_rmse_min = pred_result[2]

                # save result figure
                if self.result_fig:
                    self.save_fig(x, y, pred, i, original_result, pred_result)

                printProgressBar(i, len(self.val_data_loader),
                                 prefix="Compute measurements ..",
                                 suffix='Complete', length=25)

            l = len(self.val_data_loader)
            print('\n')
            print(
                'Original === \nPSNR avg: {:.4f} +{:.4f} -{:.4f}\nSSIM avg: {:.4f} +{:.4f} -{:.4f}\nRMSE avg: {:.4f} +{:.4f} -{:.4f}'
                .format(ori_psnr_avg / l, ori_psnr_max - ori_psnr_avg / l, ori_psnr_avg / l - ori_psnr_min,
                        ori_ssim_avg / l, ori_ssim_max - ori_ssim_avg / l, ori_ssim_avg / l - ori_ssim_min,
                        ori_rmse_avg / l, ori_rmse_max - ori_rmse_avg / l, ori_rmse_avg / l - ori_rmse_min))
            print('\n')
            print(
                'Predictions === \nPSNR avg: {:.4f} +{:.4f} -{:.4f}\nSSIM avg: {:.4f} +{:.4f} -{:.4f}\nRMSE avg: {:.4f} +{:.4f} -{:.4f}'
                .format(pred_psnr_avg / l, pred_psnr_max - pred_psnr_avg / l, pred_psnr_avg / l - pred_psnr_min,
                        pred_ssim_avg / l, pred_ssim_max - pred_ssim_avg / l, pred_ssim_avg / l - pred_ssim_min,
                        pred_rmse_avg / l, pred_rmse_max - pred_rmse_avg / l, pred_rmse_avg / l - pred_rmse_min))

            # 记录验证结果
            wb = openpyxl.load_workbook(self.model_name + '.xlsx')
            sheet = wb.get_sheet_by_name('Sheet1')
            sheet.append([pred_psnr_avg/l, pred_ssim_avg/l, pred_rmse_avg/l])
            wb.save(self.model_name + '.xlsx')


    def test(self):
       # del self.MODEL
        #self.MODEL = CTformer(img_size=64,tokens_type='performer', embed_dim=64, depth=1, num_heads=8, kernel=8, stride=4, mlp_ratio=2., token_dim=64)
        #self.CTFormer = T2T_ViT(img_size=128,tokens_type='convolution', in_chans=8,embed_dim=768, depth=6, num_heads=12, kernel=16, stride=8, mlp_ratio=2.)
        #if (self.multi_gpu) and (torch.cuda.device_count() > 1):
            #print('Use {} GPUs'.format(torch.cuda.device_count()))
            #self.MODEL = nn.DataParallel(self.MODEL)   ## data parallel

        wb = openpyxl.load_workbook(self.model_name + '.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')
        sheet.append(['TEST'])
        sheet.append(['PSNR', 'SSIM', 'RMSE'])
        wb.save(self.model_name + '.xlsx')

        self.MODEL.to(self.device)
        self.load_model_by_epoch(self.test_epoch)

        start_time = time.time()
        # compute PSNR, SSIM, RMSE
        ori_psnr_avg, ori_ssim_avg, ori_rmse_avg = 0, 0, 0
        pred_psnr_avg, pred_ssim_avg, pred_rmse_avg = 0, 0, 0

        ori_psnr_max, ori_ssim_max, ori_rmse_max = 0, 0, 0
        ori_psnr_min, ori_ssim_min, ori_rmse_min = 100, 100, 100

        pred_psnr_max, pred_ssim_max, pred_rmse_max = 0, 0, 0
        pred_psnr_min, pred_ssim_min, pred_rmse_min = 100, 100, 100

        with torch.no_grad():
            for i, (x, y) in enumerate(self.data_loader):
                # print(x.shape)
                # print(y.shape)
                shape_ = x.shape[-1]
                x = x.unsqueeze(0).float().to(self.device)
                y = y.unsqueeze(0).float().to(self.device)
                
                arrs = split_arr(x, 64).to(self.device)  ## split to image patches for test into 4 patches

                # arrs = torch.as_tensor(arrs, dtype=torch.float32)
                # print(arrs)

                arrs[0:64] = self.MODEL(arrs[0:64])
                arrs[64:2*64] = self.MODEL(arrs[64:2*64])
                arrs[2*64:3*64] = self.MODEL(arrs[2*64:3*64])
                arrs[3*64:4*64] = self.MODEL(arrs[3*64:4*64])
                pred = agg_arr(arrs, 512).to(self.device)
                

                #pred = x - pred# denormalize, truncate
                x = self.trunc(self.denormalize_(x.view(shape_, shape_).cpu().detach()))
                y = self.trunc(self.denormalize_(y.view(shape_, shape_).cpu().detach()))
                pred = self.trunc(self.denormalize_(pred.view(shape_, shape_).cpu().detach()))

                data_range = self.trunc_max - self.trunc_min

                original_result, pred_result = compute_measure(x, y, pred, data_range)
                ori_psnr_avg += original_result[0]
                if original_result[0]>ori_psnr_max:
                    ori_psnr_max = original_result[0]
                if original_result[0]<ori_psnr_min:
                    ori_psnr_min = original_result[0]

                ori_ssim_avg += original_result[1]
                if original_result[1]>ori_ssim_max:
                    ori_ssim_max = original_result[1]
                if original_result[1]<ori_ssim_min:
                    ori_ssim_min = original_result[1]

                ori_rmse_avg += original_result[2]
                if original_result[2]>ori_rmse_max:
                    ori_rmse_max = original_result[2]
                if original_result[2]<ori_rmse_min:
                    ori_rmse_min = original_result[2]

                pred_psnr_avg += pred_result[0]
                if pred_result[0]>pred_psnr_max:
                    pred_psnr_max = pred_result[0]
                if pred_result[0]<pred_psnr_min:
                    pred_psnr_min = pred_result[0]

                pred_ssim_avg += pred_result[1]
                if pred_result[1]>pred_ssim_max:
                    pred_ssim_max = pred_result[1]
                if pred_result[1]<pred_ssim_min:
                    pred_ssim_min = pred_result[1]

                pred_rmse_avg += pred_result[2]
                if pred_result[2]>pred_rmse_max:
                    pred_rmse_max = pred_result[2]
                if pred_result[2]<pred_rmse_min:
                    pred_rmse_min = pred_result[2]

                # save result figure
                if self.result_fig:
                    self.save_fig(x, y, pred, i, original_result, pred_result)

                printProgressBar(i, len(self.data_loader),
                                 prefix="Compute measurements ..",
                                 suffix='Complete', length=25)

            l = len(self.data_loader)
            print('\n')
            print('Original === \nPSNR avg: {:.4f} +{:.4f} -{:.4f}\nSSIM avg: {:.4f} +{:.4f} -{:.4f}\nRMSE avg: {:.4f} +{:.4f} -{:.4f}'
                  .format(ori_psnr_avg/l, ori_psnr_max-ori_psnr_avg/l, ori_psnr_avg/l-ori_psnr_min,
                    ori_ssim_avg/l, ori_ssim_max-ori_ssim_avg/l, ori_ssim_avg/l-ori_ssim_min,
                    ori_rmse_avg/l, ori_rmse_max-ori_rmse_avg/l, ori_rmse_avg/l-ori_rmse_min))
            print('\n')
            print('Predictions === \nPSNR avg: {:.4f} +{:.4f} -{:.4f}\nSSIM avg: {:.4f} +{:.4f} -{:.4f}\nRMSE avg: {:.4f} +{:.4f} -{:.4f}'
                  .format(pred_psnr_avg/l, pred_psnr_max-pred_psnr_avg/l, pred_psnr_avg/l-pred_psnr_min,
                  pred_ssim_avg/l, pred_ssim_max-pred_ssim_avg/l, pred_ssim_avg/l-pred_ssim_min,
                  pred_rmse_avg/l, pred_rmse_max-pred_rmse_avg/l, pred_rmse_avg/l-pred_rmse_min))

            # 记录训练总时长
            finish_time = time.time() - start_time
            print("Test time: {:.1f}s".format(finish_time))
            # 记录验证结果
            wb = openpyxl.load_workbook(self.model_name + '.xlsx')
            sheet = wb.get_sheet_by_name('Sheet1')
            sheet.append([pred_psnr_avg/l, pred_ssim_avg/l, pred_rmse_avg/l])
            wb.save(self.model_name + '.xlsx')
